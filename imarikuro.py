#!/usr/bin/env python
# coding: utf-8

# In[1]:


import pandas as pd
import numpy as np
from bs4 import BeautifulSoup
import urllib.request as req
import openpyxl
from openpyxl.styles import Font #excelのフォントの書体やサイズを操作
from openpyxl.styles.alignment import Alignment #セル内の文字の配置（左寄せ/中央/右寄せ）
from openpyxl.styles import colors #フォントカラーの操作
from openpyxl.styles import PatternFill #セルの塗りつぶし
from datetime import datetime
import pytz
import os
import shutil

# In[2]:
def scraping():
    
#取得したい天気予報のURL
    url1 = "https://tenki.jp/forecast/9/44/8520/41205/1hour.html" #Tenki.jpの伊万里市の天気
    url2 = "https://weathernews.jp/onebox/33.332689/129.842909/q=%E4%BD%90%E8%B3%80%E7%9C%8C%E4%BC%8A%E4%B8%87%E9%87%8C%E5%B8%82%E9%BB%92%E5%B7%9D%E7%94%BA%E5%A1%A9%E5%B1%8B&v=bdd3b6e1722bd658a4d39f45f58ac37a2d14d495128cd003f22d6aed1f7f81ec&temp=Temp.c&lang=ja"
#weathernewsの伊万里市黒川塩屋の天気


# In[3]:


#URLのHTMLを取得
    res1 = req.urlopen(url1)
    res2 = req.urlopen(url2)


# In[4]:


    parse1_html = BeautifulSoup(res1,'html.parser')
    parse2_html = BeautifulSoup(res2,'html.parser')


    # In[5]:


    town_h1 = parse1_html.find(id="forecast-point-1h-today")
    town_h2 = parse2_html.find(class_="switchContent__item act")


    # In[6]:


    town_h1_tm = parse1_html.find(id="forecast-point-1h-tomorrow")


    # In[7]:


    #tenki.jpの時間を取得
    hour1 = town_h1.find(class_="hour")
    h1_span = hour1.find_all("span")
    h1 = []

    for i in h1_span:
        h1.append(int(i.string)) #int型に変換して格納
        
    #過去の時間を取得
    hour1_past = hour1.find_all(class_='past')
    h1_past = []
    for i in hour1_past:
        h1_past.append(int(i.string))
        


    # In[8]:


    #tenki.jpの風速を取得
    wind_s1 = town_h1.find(class_="wind-speed")
    wind_s1_span = wind_s1.find_all("span")
    ws1 = []

    for i in wind_s1_span:
        ws1.append(int(i.string))

    #過去の風速を取得
    wind_s1_past = wind_s1.find_all(class_='past')
    ws1_past = []
    for i in wind_s1_past:
        ws1_past.append(int(i.string))
        


    # In[9]:


    #tenki.jpの風向を取得
    wind_d1 = town_h1.find(class_="wind-blow")
    wind_d1_span =wind_d1.find_all("p")
    wd1 =[]

    for i in wind_d1_span:
        wd1.append(i.string)
        
    #過去の風向を取得
    wind_d1_past = wind_d1.find_all(class_='past')
    wd1_past = []
    for i in wind_d1_past:
        wd1_past.append(i.string)


    # In[10]:


    #翌日の分も含めて24時間分にする
    #tenki.jpの時間を取得
    hour1_t = town_h1_tm.find(class_="hour")
    h1_span_t = hour1.find_all("span")
    h1_t = []

    for i in h1_span_t:
        h1.append(int(i.string)) #int型に変換して格納

    #tenki.jpの風速を取得
    wind_s1_t = town_h1_tm.find(class_="wind-speed")
    wind_s1_span_t = wind_s1_t.find_all("span")
    ws1_t = []

    for i in wind_s1_span_t:
        ws1.append(int(i.string))
        
    #tenki.jpの風向を取得
    wind_d1_t = town_h1_tm.find(class_="wind-blow")
    wind_d1_span_t =wind_d1_t.find_all("p")
    wd1_t = []

    for i in wind_d1_span_t:
        wd1.append(i.string)


    # In[11]:


    tenki_1 = pd.DataFrame({'時間':h1,'Tenki_風向':wd1,'Tenki_風速(m/s)':ws1})
    tenki_jp_past = pd.DataFrame({'時間':h1_past,'Tenki_風向':wd1_past,'Tenki_風速(m/s)':ws1_past})


    # In[12]:


    #24時間分から予報部分だけを抽出
    tenki_1 = tenki_1.merge(tenki_jp_past, indicator=True, how='outer').query('_merge=="left_only"').drop('_merge', 1)


    # In[14]:


    tenki_1 = tenki_1[:24]


    # In[16]:


    #weathernewsの時間を取得
    h2_span = town_h2.find_all(class_="wTable__item time")
    h2 = []

    for i in h2_span:
        h2.append(i.string)
    del h2[0]

    h2 = list(map(int,h2)) #int型に変換


    # In[17]:


    #weathernewsの風速を取得
    ws2_span = town_h2.find_all(class_="wTable__item w")
    ws2 = []

    for i in ws2_span:
        ws2.append(i.text)
    del ws2[0]

    #int型に変換
    ws2_int = []
    for i in ws2:
        for j in range(100):
            if i == str(j)+'m':
                ws2_int.append(j)


    # In[48]:


    #weathernewsの風向を取得
    wd2_span = town_h2.find_all(class_="wTable__item w")

    wd2_img = []
    for i in wd2_span:
        wd2_img.append(i.find('img'))

    wd2_img_src = []
    for i in wd2_img[1:]:
        wd2_img_src.append(i.attrs['src'])

    #画像データを文字に変換
    wd2 = []
    for i in wd2_img_src:
        if i == '//weathernews.jp/onebox/img/wind/wind_1_16.png' or i == '//weathernews.jp/onebox/img/wind/wind_2_16.png' or i == '//weathernews.jp/onebox/img/wind/wind_3_16.png' or i == '//weathernews.jp/onebox/img/wind/wind_4_16.png' or i == '//weathernews.jp/onebox/img/wind/wind_5_16.png':
            wd2.append('北')
        elif i =='//weathernews.jp/onebox/img/wind/wind_1_15.png' or i == '//weathernews.jp/onebox/img/wind/wind_2_15.png' or i == '//weathernews.jp/onebox/img/wind/wind_3_15.png' or i == '//weathernews.jp/onebox/img/wind/wind_4_15.png' or i == '//weathernews.jp/onebox/img/wind/wind_5_15.png':
            wd2.append('北北西')
        elif i =='//weathernews.jp/onebox/img/wind/wind_1_14.png' or i == '//weathernews.jp/onebox/img/wind/wind_2_14.png' or i == '//weathernews.jp/onebox/img/wind/wind_3_14.png' or i == '//weathernews.jp/onebox/img/wind/wind_4_14.png' or i == '//weathernews.jp/onebox/img/wind/wind_5_14.png':
            wd2.append('北西')
        elif i =='//weathernews.jp/onebox/img/wind/wind_1_13.png' or i == '//weathernews.jp/onebox/img/wind/wind_2_13.png' or i == '//weathernews.jp/onebox/img/wind/wind_3_13.png' or i == '//weathernews.jp/onebox/img/wind/wind_4_13.png' or i == '//weathernews.jp/onebox/img/wind/wind_5_13.png':
            wd2.append('西北西')
        elif i =='//weathernews.jp/onebox/img/wind/wind_1_12.png' or i == '//weathernews.jp/onebox/img/wind/wind_2_12.png' or i == '//weathernews.jp/onebox/img/wind/wind_3_12.png' or i == '//weathernews.jp/onebox/img/wind/wind_4_12.png' or i == '//weathernews.jp/onebox/img/wind/wind_5_12.png':
            wd2.append('西')
        elif i =='//weathernews.jp/onebox/img/wind/wind_1_11.png' or i == '//weathernews.jp/onebox/img/wind/wind_2_11.png' or i == '//weathernews.jp/onebox/img/wind/wind_3_11.png' or i == '//weathernews.jp/onebox/img/wind/wind_4_11.png' or i == '//weathernews.jp/onebox/img/wind/wind_5_11.png':
            wd2.append('西南西')
        elif i =='//weathernews.jp/onebox/img/wind/wind_1_10.png' or i == '//weathernews.jp/onebox/img/wind/wind_2_10.png' or i == '//weathernews.jp/onebox/img/wind/wind_3_10.png' or i == '//weathernews.jp/onebox/img/wind/wind_4_10.png' or i == '//weathernews.jp/onebox/img/wind/wind_5_10.png':
            wd2.append('南西')
        elif i =='//weathernews.jp/onebox/img/wind/wind_1_09.png' or i == '//weathernews.jp/onebox/img/wind/wind_2_09.png' or i == '//weathernews.jp/onebox/img/wind/wind_3_09.png' or i == '//weathernews.jp/onebox/img/wind/wind_4_09.png' or i == '//weathernews.jp/onebox/img/wind/wind_5_09.png':
            wd2.append('南南西')
        elif i =='//weathernews.jp/onebox/img/wind/wind_1_08.png' or i == '//weathernews.jp/onebox/img/wind/wind_2_08.png' or i == '//weathernews.jp/onebox/img/wind/wind_3_08.png' or i == '//weathernews.jp/onebox/img/wind/wind_4_08.png' or i == '//weathernews.jp/onebox/img/wind/wind_5_08.png':
            wd2.append('南')
        elif i =='//weathernews.jp/onebox/img/wind/wind_1_07.png' or i == '//weathernews.jp/onebox/img/wind/wind_2_07.png' or i == '//weathernews.jp/onebox/img/wind/wind_3_07.png' or i == '//weathernews.jp/onebox/img/wind/wind_4_07.png' or i == '//weathernews.jp/onebox/img/wind/wind_5_07.png':
            wd2.append('南南東')
        elif i =='//weathernews.jp/onebox/img/wind/wind_1_06.png' or i == '//weathernews.jp/onebox/img/wind/wind_2_06.png' or i == '//weathernews.jp/onebox/img/wind/wind_3_06.png' or i == '//weathernews.jp/onebox/img/wind/wind_4_06.png' or i == '//weathernews.jp/onebox/img/wind/wind_5_06.png':
            wd2.append('南東')
        elif i =='//weathernews.jp/onebox/img/wind/wind_1_05.png' or i == '//weathernews.jp/onebox/img/wind/wind_2_05.png' or i == '//weathernews.jp/onebox/img/wind/wind_3_05.png' or i == '//weathernews.jp/onebox/img/wind/wind_4_05.png' or i == '//weathernews.jp/onebox/img/wind/wind_5_05.png':
            wd2.append('東南東')
        elif i =='//weathernews.jp/onebox/img/wind/wind_1_04.png' or i == '//weathernews.jp/onebox/img/wind/wind_2_04.png' or i == '//weathernews.jp/onebox/img/wind/wind_3_04.png' or i == '//weathernews.jp/onebox/img/wind/wind_4_04.png' or i == '//weathernews.jp/onebox/img/wind/wind_5_04.png':
            wd2.append('東')
        elif i =='//weathernews.jp/onebox/img/wind/wind_1_03.png' or i == '//weathernews.jp/onebox/img/wind/wind_2_03.png' or i == '//weathernews.jp/onebox/img/wind/wind_3_03.png' or i == '//weathernews.jp/onebox/img/wind/wind_4_03.png' or i == '//weathernews.jp/onebox/img/wind/wind_5_03.png':
            wd2.append('東北東')
        elif i =='//weathernews.jp/onebox/img/wind/wind_1_02.png' or i == '//weathernews.jp/onebox/img/wind/wind_2_02.png' or i == '//weathernews.jp/onebox/img/wind/wind_3_02.png' or i == '//weathernews.jp/onebox/img/wind/wind_4_02.png' or i == '//weathernews.jp/onebox/img/wind/wind_5_02.png':
            wd2.append('北東')
        elif i =='//weathernews.jp/onebox/img/wind/wind_1_01.png' or i == '//weathernews.jp/onebox/img/wind/wind_2_01.png' or i == '//weathernews.jp/onebox/img/wind/wind_3_01.png' or i == '//weathernews.jp/onebox/img/wind/wind_4_01.png' or i == '//weathernews.jp/onebox/img/wind/wind_5_01.png':
            wd2.append('北北東')
        elif i =='//weathernews.jp/onebox/img/wind/wind_0_00.png':
            wd2.append('静隠')
        else: wd2.append('error')


    # In[49]:


    weather_1 = pd.DataFrame({'時間':h2,'Weather_風向':wd2,'Weather_風速(m/s)':ws2_int})


    # In[50]:
    
    weather_1 = weather_1.replace({'時間':{0:24}})
    weather_1 = weather_1[:24]


    # In[52]:


    df = pd.merge(tenki_1,weather_1,on='時間').set_index('時間')
    df = df.T
    df
    # In[24]:


    dt_now = str(datetime.now(pytz.timezone("Asia/Tokyo")).strftime('%Y_%m_%d-%H%M'))


    # In[25]:

    shutil.rmtree('./output')
    os.mkdir('./output')
    export_file = './output/風向風速予報_伊万里市_'+dt_now+'.xlsx'


    # In[26]:


    excel_sheetname = '風向風速予報'


    # In[27]:


    df.to_excel(export_file,sheet_name =excel_sheetname)


    # In[28]:


    workbook = openpyxl.load_workbook(export_file) #load_workbookでexcelファイルを読み込む
    worksheet = workbook.worksheets[0] #操作するシートの指定


    # In[29]:


    font = Font(name='メイリオ',size=14) #フォントの変更
    sheet_range = worksheet['A1':'Y7'] #フォントを変更するセルの範囲


    # In[30]:


    for row in sheet_range:
        for cell in row:
            worksheet[cell.coordinate].font = font

    worksheet.column_dimensions['A'].width = 28


    # In[31]:


    workbook.save(export_file)


    # In[32]:


    spot_name = parse2_html.find(class_='index__tit')


    # In[33]:


    worksheet['A1'].value = spot_name.string
    worksheet['A6'].value = 'お天気_風向'
    worksheet['A7'].value = 'お天気_風速(m/s)'
    worksheet['A6'].alignment = Alignment(horizontal='center')
    worksheet['A7'].alignment = Alignment(horizontal='center')


    # In[34]:


    workbook.save(export_file)
    return export_file

if __name__ == '__main__':
    scraping()




